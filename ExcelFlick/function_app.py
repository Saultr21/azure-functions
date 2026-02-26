import azure.functions as func
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)

# Índices de columna en el Excel original (1-based)
COL_DESC_VEHICULO = 3   # C - Desc vehiculo
COL_DESCRIPCION   = 5   # E - Descripción
COL_FILTRO_H      = 8   # H - P  (filtrar valor "A")
COL_DIAS          = 10  # J - Dias


@app.route(route="excel_flick")
def excel_flick(req: func.HttpRequest) -> func.HttpResponse:
    """
    Recibe un fichero Excel (.xlsx) como cuerpo binario de la petición HTTP.
    Aplica las siguientes transformaciones:
      1. Filtra las filas cuya columna H (P) sea igual a "A".
      2. Ordena por: Desc vehiculo ASC → Descripción ASC → Dias DESC.
      3. Conserva únicamente las columnas: Desc vehiculo, Descripción, Dias.
    Devuelve el Excel procesado listo para descargar.

    Uso:
        POST /api/excel_flick
        Content-Type: application/octet-stream
        Body: <bytes del fichero .xlsx>
    """
    logging.info('excel_flick: petición recibida.')

    # ------------------------------------------------------------------ #
    # 1. Leer el fichero Excel del cuerpo de la petición
    # ------------------------------------------------------------------ #
    file_bytes = req.get_body()
    if not file_bytes:
        return func.HttpResponse(
            "El cuerpo de la petición está vacío. Envía un fichero .xlsx como body.",
            status_code=400
        )

    try:
        wb_input = openpyxl.load_workbook(io.BytesIO(file_bytes))
    except Exception as exc:
        logging.error('No se pudo abrir el fichero Excel: %s', exc)
        return func.HttpResponse(
            f"No se pudo leer el fichero Excel: {exc}",
            status_code=400
        )

    ws = wb_input.active
    logging.info('Hoja activa: %s | Filas: %d | Columnas: %d',
                 ws.title, ws.max_row, ws.max_column)

    # ------------------------------------------------------------------ #
    # 2. Extraer cabecera y filas de datos
    # ------------------------------------------------------------------ #
    header_row = [cell.value for cell in ws[1]]

    # Nombre limpio de cabeceras (strip de espacios)
    def col_name(col_idx: int) -> str:
        """Devuelve el nombre de cabecera (strip) para el índice 1-based."""
        return str(header_row[col_idx - 1]).strip() if header_row[col_idx - 1] else ""

    # ------------------------------------------------------------------ #
    # 3. Filtrar: col H == "A"
    # ------------------------------------------------------------------ #
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        val_h = row[COL_FILTRO_H - 1]  # índice 0-based
        if val_h == "A":
            data_rows.append(row)

    logging.info('Filas tras filtro H=A: %d', len(data_rows))

    # ------------------------------------------------------------------ #
    # 4. Ordenar: Desc vehiculo ASC, Descripción ASC, Dias DESC
    # ------------------------------------------------------------------ #
    def sort_key(row):
        desc_vehiculo = row[COL_DESC_VEHICULO - 1] or ""
        descripcion   = row[COL_DESCRIPCION - 1]   or ""
        dias          = row[COL_DIAS - 1]           or 0
        return (str(desc_vehiculo).strip().lower(),
                str(descripcion).strip().lower(),
                -int(dias) if isinstance(dias, (int, float)) else 0)

    data_rows.sort(key=sort_key)
    logging.info('Filas ordenadas correctamente.')

    # ------------------------------------------------------------------ #
    # 5. Construir el nuevo workbook con solo las 3 columnas necesarias
    # ------------------------------------------------------------------ #
    wb_output = openpyxl.Workbook()
    ws_out = wb_output.active
    ws_out.title = "VTO Procesado"

    # Cabeceras del resultado
    output_headers = [
        col_name(COL_DESC_VEHICULO),  # Desc vehiculo
        col_name(COL_DESCRIPCION),    # Descripción
        col_name(COL_DIAS),           # Dias
    ]

    # Estilo de cabecera
    header_font    = Font(bold=True, color="FFFFFF")
    header_fill    = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align   = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    ws_out.append(output_headers)
    for col_idx, _ in enumerate(output_headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    # Filas de datos
    data_align = Alignment(horizontal="left", vertical="center")
    for row in data_rows:
        output_row = [
            row[COL_DESC_VEHICULO - 1],
            row[COL_DESCRIPCION - 1],
            row[COL_DIAS - 1],
        ]
        ws_out.append(output_row)
        row_idx = ws_out.max_row
        for col_idx in range(1, 4):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            cell.alignment = data_align
            cell.border    = thin_border

    # Ajustar ancho de columnas automáticamente
    for col_idx in range(1, 4):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(output_headers[col_idx - 1]))
        for row_idx in range(2, ws_out.max_row + 1):
            val = ws_out.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        ws_out.column_dimensions[col_letter].width = min(max_length + 4, 60)

    # Fijar fila de cabecera
    ws_out.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # 6. Devolver el fichero procesado
    # ------------------------------------------------------------------ #
    output_buffer = io.BytesIO()
    wb_output.save(output_buffer)
    output_buffer.seek(0)

    logging.info('Fichero procesado generado con %d filas de datos.', len(data_rows))

    return func.HttpResponse(
        body=output_buffer.getvalue(),
        status_code=200,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=\"VTO_Procesado.xlsx\""
        }
    )