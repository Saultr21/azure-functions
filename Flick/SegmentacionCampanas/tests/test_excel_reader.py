import io
from datetime import date

import openpyxl
import pytest

from excel_reader import leer_registros, ColumnaFaltanteError

COLUMNAS = [
    "Nº.matrícula", "Descripción", "Direccion(3)", "Fecha.matriculación",
    "Fecha.de.servicio", "Kilometraje", "Kilometraje.mantenim",
    "Código.de.servicio(1)", "Fin.mantenimiento", "Fecha.exp.garantia",
    "Inicio.garant.extend", "Saludo.(tratamiento)", "Nº.telefono(4)",
    "Direccion.e-mail",
]


def _crear_excel_bytes(filas: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(COLUMNAS)
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_lee_una_fila_correctamente():
    fila = [
        "1234ABC", "Yamaha NMAX 125", "Telde", date(2020, 1, 1),
        date(2024, 1, 1), 5000, 10000, "OK1", "--/--/--", "--/--/--",
        "--/--/--", "Sr.", "600123456", "cliente@example.com",
    ]
    excel_bytes = _crear_excel_bytes([fila])

    registros = leer_registros(excel_bytes)

    assert len(registros) == 1
    assert registros[0].matricula == "1234ABC"
    assert registros[0].fecha_matriculacion == date(2020, 1, 1)
    assert registros[0].km_ultimo_servicio == 5000


def test_fila_sin_matricula_se_descarta():
    fila = ["", "Yamaha NMAX 125", "Telde", date(2020, 1, 1), date(2024, 1, 1),
            5000, 10000, "OK1", "--/--/--", "--/--/--", "--/--/--", "Sr.",
            "600123456", "cliente@example.com"]
    excel_bytes = _crear_excel_bytes([fila])

    registros = leer_registros(excel_bytes)

    assert registros == []


def test_falta_una_columna_obligatoria_lanza_error():
    wb_bytes_sin_matricula = _crear_excel_bytes([])
    # Reescribimos quitando la cabecera de matrícula
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes_sin_matricula))
    ws = wb["Hoja1"]
    ws.cell(row=1, column=1).value = "OtraColumna"
    buffer = io.BytesIO()
    wb.save(buffer)

    with pytest.raises(ColumnaFaltanteError):
        leer_registros(buffer.getvalue())
