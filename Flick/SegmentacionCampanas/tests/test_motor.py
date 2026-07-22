from datetime import date
import io

import openpyxl
import pytest

from models import CampanaId
from campanas.motor import ejecutar_campana, CampanaNoSoportadaError

COLUMNAS = [
    "Nº.matrícula", "Descripción", "Direccion(3)", "Fecha.matriculación",
    "Fecha.de.servicio", "Kilometraje", "Kilometraje.mantenim",
    "Código.de.servicio(1)", "Fin.mantenimiento", "Fecha.exp.garantia",
    "Inicio.garant.extend", "Saludo.(tratamiento)", "Nº.telefono(4)",
    "Direccion.e-mail",
]


def _excel_con_una_fila_valida_3m() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(COLUMNAS)
    ws.append([
        "1234ABC", "Yamaha NMAX 125", "Telde", date(2023, 1, 1),
        date(2020, 1, 1), 50, 10000, "OK1", "--/--/--", "--/--/--",
        "--/--/--", "Sr.", "600123456", "cliente@example.com",
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_ejecutar_campana_3m_devuelve_registros_csv_y_excel():
    excel_bytes = _excel_con_una_fila_valida_3m()

    resultado = ejecutar_campana(CampanaId.TRES_MESES, excel_bytes, hoy=date(2026, 7, 21))

    assert resultado.total_clientes == 1
    assert "1234ABC" in resultado.csv_contenido
    assert resultado.nombre_archivo == "FiltradoCampana3M_2026-07-21.xlsx"

    libro = openpyxl.load_workbook(io.BytesIO(resultado.excel_contenido))
    valores_fila = [celda.value for celda in libro.active[2]]
    assert "1234ABC" in valores_fila

    # Ningún municipio no reconocido en este Excel de prueba (Telde es válido).
    assert resultado.municipios_no_reconocidos == {}


def test_ejecutar_campana_reporta_municipios_no_reconocidos():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(COLUMNAS)
    # Modelo Yamaha válido pero municipio fuera de la lista -- se descarta
    # del resultado, pero debe contarse como municipio no reconocido.
    ws.append([
        "9999XYZ", "Yamaha NMAX 125", "Agaete", date(2023, 1, 1),
        date(2020, 1, 1), 50, 10000, "OK1", "--/--/--", "--/--/--",
        "--/--/--", "Sr.", "600999999", "otro@example.com",
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()

    resultado = ejecutar_campana(CampanaId.TRES_MESES, excel_bytes, hoy=date(2026, 7, 21))

    assert resultado.total_clientes == 0
    assert resultado.municipios_no_reconocidos == {"Agaete": 1}


def test_campana_no_soportada_lanza_error():
    with pytest.raises(CampanaNoSoportadaError):
        ejecutar_campana("NO-EXISTE", b"", hoy=date(2026, 7, 21))
