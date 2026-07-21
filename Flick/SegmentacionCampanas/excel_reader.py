"""Lector del Excel maestro de clientes.

Localiza columnas por coincidencia normalizada de cabecera (igual que `col()`
en los Office Scripts originales), no por posición fija, porque las cabeceras
reales tienen variaciones de nombre y sufijos numéricos (p. ej. "Nº.telefono(4)").
"""

import io
import re
from typing import Optional

import openpyxl

from models import RegistroCliente
from utils import normalizar_texto, parsear_fecha

HOJA_ESPERADA = "Hoja1"

# nombre_campo -> fragmento de cabecera a buscar (normalizado, "contains")
COLUMNAS_REQUERIDAS = {
    "matricula": "matricula",
    "descripcion": "descripcion",
    "municipio": "direccion(3)",
    "fecha_matriculacion": "fecha.matriculacion",
    "fecha_servicio": "fecha.de.servicio",
    "codigo_servicio": "codigo.de.servicio",
    "fin_mantenimiento": "fin.mantenimiento",
    "km_mantenimiento": "kilometraje.mantenim",
    "fecha_exp_garantia": "fecha.exp.garantia",
    "inicio_garantia_extendida": "inicio.garant.extend",
    "saludo": "saludo",
    "telefono": "telefono",
    "email": "e-mail",
}
# Excepción: "Kilometraje" (a secas) hay que matchearlo por igualdad exacta,
# porque "contains" también encontraría "Kilometraje.mantenim" (igual que en
# los .osts, que usan colExact para esta columna).
COLUMNA_KM_ULTIMO_SERVICIO = "kilometraje"


class ColumnaFaltanteError(Exception):
    pass


def _token_coincide(token_cabecera: str, token_fragmento: str) -> bool:
    """Un token de cabecera coincide con un token de fragmento si son iguales,
    o si el token de cabecera es el fragmento seguido de un sufijo numérico
    entre paréntesis (p. ej. "servicio(1)" para el fragmento "servicio").

    No se usa "in" (substring) porque eso hace que fragmentos cortos como
    "matricula" encajen dentro de palabras distintas más largas como
    "matriculacion" (de "Fecha.matriculación"), produciendo falsos positivos.
    """
    if token_cabecera == token_fragmento:
        return True
    return re.fullmatch(re.escape(token_fragmento) + r"\(\d+\)", token_cabecera) is not None


def _fragmento_coincide(tokens_cabecera: list[str], tokens_fragmento: list[str]) -> bool:
    n, m = len(tokens_cabecera), len(tokens_fragmento)
    for inicio in range(n - m + 1):
        if all(
            _token_coincide(tokens_cabecera[inicio + i], tokens_fragmento[i])
            for i in range(m)
        ):
            return True
    return False


def _construir_indice_columnas(cabeceras: list[str]) -> dict[str, int]:
    cabeceras_norm = [normalizar_texto(h) for h in cabeceras]
    tokens_por_cabecera = [h.split(".") for h in cabeceras_norm]
    indice: dict[str, int] = {}

    for campo, fragmento in COLUMNAS_REQUERIDAS.items():
        tokens_fragmento = fragmento.split(".")
        encontrada = next(
            (
                i for i, tokens in enumerate(tokens_por_cabecera)
                if _fragmento_coincide(tokens, tokens_fragmento)
            ),
            None,
        )
        if encontrada is None:
            raise ColumnaFaltanteError(f"No se encontró la columna para '{campo}' ({fragmento})")
        indice[campo] = encontrada

    idx_km = next((i for i, h in enumerate(cabeceras_norm) if h == COLUMNA_KM_ULTIMO_SERVICIO), None)
    if idx_km is None:
        raise ColumnaFaltanteError("No se encontró la columna exacta 'Kilometraje'")
    indice["km_ultimo_servicio"] = idx_km

    return indice


def leer_registros(excel_bytes: bytes) -> list[RegistroCliente]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=True)
    if HOJA_ESPERADA not in wb.sheetnames:
        raise ColumnaFaltanteError(f"No existe la hoja '{HOJA_ESPERADA}'")
    hoja = wb[HOJA_ESPERADA]

    filas = hoja.iter_rows(values_only=True)
    cabeceras = [str(c) if c is not None else "" for c in next(filas)]
    indice = _construir_indice_columnas(cabeceras)

    registros: list[RegistroCliente] = []
    for fila in filas:
        matricula = str(fila[indice["matricula"]] or "").strip()
        if not matricula:
            continue

        registros.append(
            RegistroCliente(
                matricula=matricula,
                descripcion=str(fila[indice["descripcion"]] or ""),
                municipio=str(fila[indice["municipio"]] or ""),
                fecha_matriculacion=parsear_fecha(fila[indice["fecha_matriculacion"]]),
                fecha_servicio=parsear_fecha(fila[indice["fecha_servicio"]]),
                km_ultimo_servicio=_a_numero(fila[indice["km_ultimo_servicio"]]),
                km_mantenimiento=_a_numero(fila[indice["km_mantenimiento"]]),
                codigo_servicio=str(fila[indice["codigo_servicio"]] or "").strip(),
                fin_mantenimiento=parsear_fecha(fila[indice["fin_mantenimiento"]]),
                fecha_exp_garantia=parsear_fecha(fila[indice["fecha_exp_garantia"]]),
                inicio_garantia_extendida=parsear_fecha(fila[indice["inicio_garantia_extendida"]]),
                saludo=str(fila[indice["saludo"]] or ""),
                telefono=str(fila[indice["telefono"]] or ""),
                email=str(fila[indice["email"]] or ""),
            )
        )

    return registros


def _a_numero(valor) -> Optional[float]:
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None
